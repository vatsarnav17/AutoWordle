from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill
import os

GREEN = PatternFill(start_color="00C853",end_color="00C853",fill_type="solid")
YELLOW = PatternFill(start_color="FFD600",end_color="FFD600",fill_type="solid")
GREY = PatternFill(start_color="BDBDBD",end_color="BDBDBD",fill_type="solid")

class Excel_Logger:
    def __init__(self,path="wordle log.xlsx"):
        self.path=path

        if os.path.exists(self.path):
            self.wb = load_workbook(self.path)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.append(["Game","Turn","C1","C2","C3","C4","C5"])
            self.wb.save(self.path)
        
        # Get the last game number from the Game column
        last_game = 0
        for row in range(2, self.ws.max_row + 1):  # Start from row 2 (skip header)
            cell_value = self.ws.cell(row=row, column=1).value
            if cell_value is not None:
                last_game = max(last_game, int(cell_value))
        
        self.game_id = last_game

    def start_new_game(self):
        # Add a blank row for spacing between games (only if not the first game)
        if self.game_id > 0:
            self.ws.append([])
        self.game_id+=1

    def log_turn(self,turn,feedback):
        row=[self.game_id,turn+1]

        for letter,_ in feedback:
            if letter is not None:
                row.append(letter.upper())

        self.ws.append(row)
        r = self.ws.max_row

        for i,(_,evaluation) in enumerate(feedback,start =3):
            cell = self.ws.cell(row=r,column=i)

            if evaluation =="correct":
                cell.fill = GREEN
            elif evaluation =="present":
                cell.fill = YELLOW
            elif evaluation =="absent":
                cell.fill = GREY
        
        self.wb.save(self.path)

    exceptions = {
        "genius",
        "magnificent",
        "impressive",
        "splendid",
        "great",
        "phew"
    }
    
    def log_answer(self,answer):
        ans = answer.lower()

        if ans not in self.exceptions:
            self.ws.append([self.game_id,"ANSWER",answer.upper()])
            self.wb.save(self.path)

